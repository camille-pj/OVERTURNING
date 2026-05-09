"""
Girder Overturning Check
========================
Replicates GIRDER_OVERTURNING_CHECK.xlsx (Sheet1) for a launching girder on a
pin support: back span on the left holds the system down, the right girder span
plus the launching nose at the right tip try to tip it over.

All numerical parameters are read from the workbook so the script tracks any
edits to the source file. Only L_left is supplied at runtime.

Source-cell map (Sheet1) -- every parameter traces back to a cell:

    D31  A_girder_mm2     girder cross-sectional area                 (mm^2)
    D32  rho_concrete     concrete unit weight                        (kN/m^3)
    D34  w_girder_per_m   girder weight per metre = D31*D32/1e6       (kN/m)
    D36  W_LN             total launching-nose weight (flat input)    (kN)
    D37  arm_LN_from_end  nose-centroid arm from girder end (= L_LN/3 in source) (m)
    D49  L_total          total system length (girder + nose)         (m)
    F65  SF_required      required overturning safety factor          (-)

Workbook formulas reproduced verbatim:

    H56  L_right        = L_total - L_left
    D57  W_left         = L_left  * w_girder_per_m
    H57  W_right        = L_right * w_girder_per_m
    H59  LN_moment      = W_LN    * (L_right + arm_LN_from_end)
    D62  RM             = W_left  * L_left  / 2
    H62  OM             = W_right * L_right / 2 + LN_moment
    D68  SF             = RM / OM
    D72  Deficit        = OM * SF_required - RM

NOTE on geometry: the workbook defines L_total = 25.145 + 8 = 33.145 m, i.e.
girder concrete (25.145 m) + nose (8 m). H56 then sets L_right = L_total - L_left,
so L_right spans into the nose region and W_right is mass*length at girder unit
weight over that full length, while the nose weight is added separately via H59.
That is the source workbook's modelling choice and is reproduced here as-is.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

# Make stdout UTF-8 so the formatted summary (kN-m, check marks) prints cleanly
# on Windows consoles that default to cp1252.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, OSError):
    pass

XLSX_PATH = Path(r"C:\Users\Camille\Downloads\GIRDER OVERTURNING CHECK (1).xlsx")


def load_inputs(path: Path) -> dict:
    """Read every parameter from its named cell. data_only=True returns the
    cached calculated value, so each cell's formula must have been evaluated
    by Excel and saved at least once."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active  # single sheet "Sheet1"

    A_girder_mm2    = ws["D31"].value   # mm^2
    rho_concrete    = ws["D32"].value   # kN/m^3
    w_girder_per_m  = ws["D34"].value   # kN/m  (= rho*A/1e6)
    W_LN            = ws["D36"].value   # kN    (total nose weight)
    arm_LN_from_end = ws["D37"].value   # m     (nose centroid from girder end)
    L_total         = ws["D49"].value   # m     (girder + nose)
    SF_required     = ws["F65"].value   # -

    sf_missing = SF_required is None
    if sf_missing:
        SF_required = 1.5

    # Sanity: ensure the cached per-metre weight matches rho*A/1e6 exactly.
    expected_w = rho_concrete * A_girder_mm2 / 1_000_000.0
    if abs(expected_w - w_girder_per_m) > 1e-9:
        print(
            f"Warning: D34 ({w_girder_per_m}) != D32*D31/1e6 ({expected_w:.6f}); "
            "using D34 as authoritative."
        )

    return dict(
        A_girder_mm2=A_girder_mm2,
        rho_concrete=rho_concrete,
        w_girder_per_m=w_girder_per_m,
        W_LN=W_LN,
        arm_LN_from_end=arm_LN_from_end,
        L_total=L_total,
        SF_required=SF_required,
        sf_missing=sf_missing,
    )


def compute(L_left: float, p: dict) -> dict:
    L_total         = p["L_total"]
    w_per_m         = p["w_girder_per_m"]
    W_LN            = p["W_LN"]
    arm_LN_from_end = p["arm_LN_from_end"]
    SF_required     = p["SF_required"]

    L_right   = L_total - L_left                 # H56
    W_left    = L_left  * w_per_m                # D57
    W_right   = L_right * w_per_m                # H57

    arm_left  = L_left  / 2.0                    # D62 implicit
    arm_right = L_right / 2.0                    # H62 implicit
    arm_LN    = L_right + arm_LN_from_end        # H59 implicit

    M_stab_left   = W_left  * arm_left           # D62
    M_over_right  = W_right * arm_right
    M_over_LN     = W_LN    * arm_LN             # H59

    M_stabilizing = M_stab_left
    M_overturning = M_over_right + M_over_LN     # H62

    SF      = M_stabilizing / M_overturning      # D68
    deficit = M_overturning * SF_required - M_stabilizing  # D72

    return dict(
        L_right=L_right,
        W_left=W_left, W_right=W_right, W_LN=W_LN,
        arm_left=arm_left, arm_right=arm_right, arm_LN=arm_LN,
        M_stabilizing=M_stabilizing, M_overturning=M_overturning,
        SF=SF, SF_required=SF_required, deficit=deficit,
    )


def print_report(L_left: float, r: dict) -> None:
    pass_ = r["SF"] >= r["SF_required"]
    status = "✅ PASS" if pass_ else "❌ FAIL"
    deficit_label = "(excess capacity)" if r["deficit"] < 0 else "(moment deficit)"

    bar = "=" * 60
    print()
    print(bar)
    print("  GIRDER OVERTURNING CHECK")
    print("  Developed by Albert Pamonag and Camille Pajarillaga")
    print(bar)
    print("  Input")
    print(f"    L_left              : {L_left:.2f} m")
    print()
    print("  Loads")
    print(f"    W_left              : {r['W_left']:.2f} kN  @ {r['arm_left']:.2f} m from pin")
    print(f"    W_right             : {r['W_right']:.2f} kN  @ {r['arm_right']:.2f} m from pin")
    print(f"    W_LN                : {r['W_LN']:.2f} kN  @ {r['arm_LN']:.2f} m from pin")
    print()
    print("  Moments about pin")
    print(f"    Stabilizing (left)  : {r['M_stabilizing']:.2f} kN·m")
    print(f"    Overturning (right) : {r['M_overturning']:.2f} kN·m")
    print()
    print("  Result")
    print(f"    Safety Factor       : {r['SF']:.3f}")
    print(f"    Required SF         : {r['SF_required']:.2f}")
    print(f"    Status              : {status}")
    print(f"    Deficit             : {abs(r['deficit']):.2f} kN·m {deficit_label}")
    print(bar)


def main() -> None:
    p = load_inputs(XLSX_PATH)
    if p["sf_missing"]:
        print("Warning: SF_required not found in F65 — defaulting to 1.5")

    L_left = float(input("Enter left girder length L_left (m): "))

    if L_left <= 0 or L_left >= p["L_total"]:
        raise SystemExit(
            f"L_left must be in (0, {p['L_total']:.3f}) m to keep the system on the pin."
        )

    r = compute(L_left, p)
    print_report(L_left, r)


if __name__ == "__main__":
    main()
